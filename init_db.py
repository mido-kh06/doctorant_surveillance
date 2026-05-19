import sqlite3
import os
import openpyxl

def init_db():
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')
    if os.path.exists(db_path):
        os.remove(db_path)
        
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Students table - pre-loaded from Excel
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        anne_inscription TEXT,
        nins TEXT,
        nom_complet TEXT NOT NULL,
        prenom TEXT,
        cne TEXT,
        cin TEXT NOT NULL UNIQUE,
        labo TEXT,
        situation TEXT,
        email TEXT,
        chosen_day TEXT DEFAULT NULL,
        assigned_period TEXT DEFAULT NULL,
        email_sent INTEGER DEFAULT 0,
        registration_date TIMESTAMP DEFAULT NULL
    )
    ''')
    
    # Days table with capacity
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        day_label TEXT NOT NULL UNIQUE,
        day_date TEXT NOT NULL,
        max_capacity INTEGER DEFAULT 20,
        is_active INTEGER DEFAULT 1
    )
    ''')
    
    # Settings table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    ''')
    
    # Insert exam days
    exam_days = [
        ("الإثنين 08 يونيو 2026", "2026-06-08"),
        ("الثلاثاء 09 يونيو 2026", "2026-06-09"),
        ("الأربعاء 10 يونيو 2026", "2026-06-10"),
        ("الخميس 11 يونيو 2026", "2026-06-11"),
        ("الجمعة 12 يونيو 2026", "2026-06-12"),
        ("الإثنين 15 يونيو 2026", "2026-06-15"),
        ("الثلاثاء 16 يونيو 2026", "2026-06-16"),
    ]
    
    for label, date in exam_days:
        cursor.execute('INSERT INTO days (day_label, day_date, max_capacity) VALUES (?, ?, 20)', (label, date))
    
    # Admin settings
    settings = {
        'admin_user': 'memdouh',
        'admin_password': 'memdouh123',
        'smtp_host': 'smtp.gmail.com',
        'smtp_port': '587',
        'smtp_user': '',
        'smtp_password': '',
        'email_subject': 'استدعاء حراسة امتحانات الدورة الربيعية - كلية الشريعة بفاس',
        'email_template': 'مرحباً {nom_complet}،\n\nتم تأكيد موعد حراستكم لامتحانات الدورة الربيعية العادية 2025/2026.\n\nتفاصيل الحراسة:\n- اليوم: {day_label}\n- الفترة: {period}\n\nمعلومات الطالب:\n- رقم CIN: {cin}\n- رقم التسجيل: {nins}\n- رمز CNE: {cne}\n\nيرجى الحضور في الوقت المحدد.\n\nمع تحيات كلية الشريعة، جامعة سيدي محمد بن عبد الله، فاس.'
    }
    
    for key, value in settings.items():
        cursor.execute("INSERT INTO settings (key, value) VALUES (?, ?)", (key, value))
    
    # Load students from Excel
    excel_path = r'C:\Users\yassm\Downloads\mehdi.xlsx'
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        count = 0
        skipped = 0
        for row_idx in range(2, ws.max_row + 1):
            row = [cell.value for cell in ws[row_idx]]
            
            # Headers: anne inscription, nins, anom, PrénomAR, CNE, CIN, Labo, الوضعية, Email
            anne = str(row[0]) if row[0] else ''
            nins = str(row[1]) if row[1] else ''
            nom = str(row[2]) if row[2] else ''
            prenom = str(row[3]) if row[3] else ''
            cne = str(row[4]) if row[4] else ''
            cin = str(row[5]).strip().upper() if row[5] else ''
            labo = str(row[6]) if row[6] else ''
            situation = str(row[7]) if row[7] else ''
            email = str(row[8]).strip().lower() if row[8] else ''
            
            if not cin or cin == 'None' or cin == '':
                skipped += 1
                continue
                
            try:
                cursor.execute('''
                    INSERT OR IGNORE INTO students 
                    (anne_inscription, nins, nom_complet, prenom, cne, cin, labo, situation, email)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (anne, nins, nom, prenom, cne, cin, labo, situation, email))
                if cursor.rowcount > 0:
                    count += 1
                else:
                    skipped += 1
            except Exception as e:
                skipped += 1
                print(f"  Skipped row {row_idx}: {e}")
                
        print(f"Loaded {count} students from Excel. Skipped {skipped} duplicates/empty.")
    else:
        print(f"Excel file not found at: {excel_path}")
    
    conn.commit()
    conn.close()
    print("Database initialized successfully!")

if __name__ == '__main__':
    init_db()
